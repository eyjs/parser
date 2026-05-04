"""Excel parser adapter — Excel bytes (.xlsx/.xls) to Markdown table conversion.

Ported from ai-platform's excel_parser.py, adapted to DocForge conventions:
frozen dataclasses, pure functions, logging module.

Requires openpyxl (optional dependency: ``pip install docforge[excel]``).
"""

from __future__ import annotations

import io
import logging
import time
from dataclasses import dataclass

logger = logging.getLogger(__name__)

_MAX_ROWS_PER_SHEET_DEFAULT = 10000
_MAX_SHEETS_DEFAULT = 20


@dataclass(frozen=True)
class ExcelParseResult:
    """Excel 파싱 결과."""

    markdown: str
    metadata: dict
    stats: dict


def parse_excel_bytes(
    data: bytes,
    filename: str = "",
    max_rows_per_sheet: int = _MAX_ROWS_PER_SHEET_DEFAULT,
    max_sheets: int = _MAX_SHEETS_DEFAULT,
) -> ExcelParseResult:
    """Excel 바이트 데이터를 Markdown 테이블로 변환한다.

    Parameters
    ----------
    data:
        .xlsx 또는 .xls 파일의 바이트 데이터.
    filename:
        원본 파일명 (로깅용).
    max_rows_per_sheet:
        시트당 최대 행 수.
    max_sheets:
        최대 처리 시트 수.

    Returns
    -------
    ExcelParseResult:
        markdown, metadata, stats를 포함하는 불변 결과.

    Raises
    ------
    ImportError:
        openpyxl이 설치되지 않은 경우.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel parsing: pip install docforge[excel]"
        )

    t0 = time.time()

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    sections: list[str] = []
    total_rows = 0
    sheet_count = 0

    for sheet_name in wb.sheetnames[:max_sheets]:
        ws = wb[sheet_name]
        rows: list[list[str]] = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows_per_sheet:
                break
            cells = [str(c) if c is not None else "" for c in row]
            # skip completely empty rows
            if any(c.strip() for c in cells):
                rows.append(cells)

        if not rows:
            continue

        sheet_count += 1
        total_rows += len(rows)
        section = _sheet_to_markdown(sheet_name, rows)
        sections.append(section)

    wb.close()

    markdown = "\n\n".join(sections)
    elapsed_ms = (time.time() - t0) * 1000

    logger.info(
        "excel_parsed: filename=%s, sheets=%d, total_rows=%d, latency_ms=%.1f",
        filename,
        sheet_count,
        total_rows,
        elapsed_ms,
    )

    return ExcelParseResult(
        markdown=markdown,
        metadata={
            "filename": filename,
            "sheets": sheet_count,
            "total_rows": total_rows,
        },
        stats={"parse_time_ms": round(elapsed_ms, 2)},
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _sheet_to_markdown(sheet_name: str, rows: list[list[str]]) -> str:
    """시트 데이터를 Markdown 테이블 섹션으로 변환한다."""
    if not rows:
        return ""

    header = rows[0]
    data_rows = rows[1:]
    col_count = len(header)

    parts: list[str] = [f"## {sheet_name}\n"]

    # auto-name blank header cells
    header_cells: list[str] = []
    for i, h in enumerate(header):
        cell = h.strip()
        if not cell:
            cell = f"Column_{i + 1}"
        header_cells.append(_escape_pipe(cell))

    parts.append("| " + " | ".join(header_cells) + " |")
    parts.append("| " + " | ".join(["---"] * col_count) + " |")

    for row in data_rows:
        # normalize column count
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        elif len(row) > col_count:
            row = row[:col_count]
        cells = [_escape_pipe(c) for c in row]
        parts.append("| " + " | ".join(cells) + " |")

    return "\n".join(parts)


def _escape_pipe(value: str) -> str:
    """마크다운 테이블 셀 이스케이프."""
    return value.replace("|", "\\|").replace("\n", " ").strip()
